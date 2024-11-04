# 套件匯入
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pyautogui as pag
from selenium.webdriver.chrome.service import Service
import requests
from http import HTTPStatus
from selenium.webdriver.chrome.options import Options
from lxml import html
import csv
import numpy as np
import pandas as pd
import cv2
import pybi as pbi
import os
import sys
import aspose
import xlrd
import xlwt
import glob
import cx_Freeze
from cx_Freeze import setup
import setuptools
import jpype
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType
import difflib
import openpyxl
from openpyxl import Workbook
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl.styles import Font  # 導入字體模組
from openpyxl.styles import PatternFill  # 導入填充模組
from spire.xls import *
from collections import deque
from pandas.core.frame import DataFrame
# 导入openpyxl模块并将其重命名为pxl
import openpyxl as pxl
# 从openpyxl导入PatternFill类
from openpyxl.styles import PatternFill
# 从openpyxl导入colors类
from openpyxl.styles import colors
# 从openpyxl导入Font类
from openpyxl.styles import Font
import datetime
import pytesseract
from PIL import Image
import ddddocr
import logging


current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
print('測試起始時間: ', current_time, '\n')

print('串接(菲律賓骰寶)前台API擷取遊戲紀錄測試中...', '\n')
time.sleep(1)

# ch_options = Options()
# ch_options.add_argument("--headless")  # 無WEB UI顯示
# driver = webdriver.Chrome(ch_options)
   
driver = webdriver.Chrome() 
driver.get('https://dev-api.iplaystar.net/gamehistory/?host_id=aa62ffb88b300f6be6654615f17ce6fa&lang=tch&game_id=PSC-ON-00016&count=20&page=1&uid=uN22JyloR1aD2zgUIu8nL1ogWFBPqrujwHHpugHyw94%3d')
http_status = requests.get('https://dev-api.iplaystar.net/gamehistory/?host_id=aa62ffb88b300f6be6654615f17ce6fa&lang=tch&game_id=PSC-ON-00016&count=20&page=1&uid=uN22JyloR1aD2zgUIu8nL1ogWFBPqrujwHHpugHyw94%3d')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print("HTTP回應失敗!", '\n')    
    print('自動化測試已中斷!', '\n')
    time.sleep(2)
    logging.basicConfig(level=logging.DEBUG,
                    filename='output.log',
                    datefmt='%Y/%m/%d %H:%M:%S',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(lineno)d - %(module)s - %(message)s')
    logger = logging.getLogger(__name__)
    logger.info('This is a log info')
    logger.debug('Debugging')
    logger.warning('Warning exists')
    logger.info('Finish')
    driver.quit()


# =================================選取記錄日期 =================================

dateList = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="serdate"]')))  # 點選日期選單
actions = ActionChains(driver)
actions.move_to_element(dateList)    
actions.perform()
dateList.click()
time.sleep(2)

recordDate = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/thead/tr[2]/th[1]')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(recordDate)    
actions.perform()
recordDate.click()

date_select01 = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/tbody/tr[3]/td[6]')  # 點選日期(2024.09.13)
date_select01.click()
time.sleep(1)


# # =================================選取遊戲名稱 =================================

# gameList = WebDriverWait(driver, 5).until(
#     EC.element_to_be_clickable((By.XPATH, '//*[@id="game_select"]')))  # 點選遊戲選單
# actions = ActionChains(driver)
# actions.move_to_element(gameList)    
# actions.perform()
# gameList.click()
# time.sleep(2)

# # game_select = driver.find_element(By.XPATH, '//*[@id="game_select"]/option[8]')  # 點選遊戲名稱(PSC-ON-00016_菲律賓骰寶)
# # game_select.click()
# # time.sleep(2)
print('獲取(菲律賓骰寶)前台遊戲紀錄中...', '\n')

# =======================================================寫入第一筆測試紀錄=======================================================

serialNum_1 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[2]/td')
numText1 = serialNum_1.text
numList1 = list(numText1)
serialNumlist = DataFrame(numList1)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏'])

xpath1 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[1]/td/div')  # 房間底注 / 籌碼 / 總輸贏 元素位置
textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newList1.remove('房間底注:')
newList1.remove('籌碼:')
newList1.remove('總輸贏:')     
            
newDataList1 = DataFrame(newList1)
newList1T = newDataList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')
    
for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[1]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    # print(textSplit2, '\n')
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile2 = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="content"]/div[1]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[1]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第二筆測試紀錄=======================================================

serialNum_2 = driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/table/tbody/tr[2]/td')
numText2 = serialNum_2.text
numList2 = list(numText2)
serialNumlist = DataFrame(numList2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏'])
 
xpath2 = driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/table/tbody/tr[1]/td/div')
textSplit2 = xpath2.text.split()
newList2 = list(textSplit2)
# print(newList1, '\n')
newList2.remove('房間底注:')
newList2.remove('籌碼:')
newList2.remove('總輸贏:')
# print(newList1, '\n')
newDataList2 = DataFrame(newList2)
newList2T = newDataList2.T
csvFile = newList2T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[2]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    # print(textSplit2, '\n')
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile2 = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[2]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[2]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第三筆測試紀錄=======================================================

serialNum_3 = driver.find_element(By.XPATH, '//*[@id="content"]/div[3]/div/table/tbody/tr[2]/td')
numText3 = serialNum_3.text
numList3 = list(numText3)
serialNumlist = DataFrame(numList3)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath3 = driver.find_element(By.XPATH, '//*[@id="content"]/div[3]/div/table/tbody/tr[1]/td/div')
textSplit3 = xpath3.text.split()
newList3 = list(textSplit3)
# print(newList1, '\n')
newList3.remove('房間底注:')
newList3.remove('籌碼:')
newList3.remove('總輸贏:')
# print(newList1, '\n')
newDataList3 = DataFrame(newList3)
newList3T = newDataList3.T
csvFile = newList3T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[3]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    # print(textSplit2, '\n')
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile2 = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[3]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[3]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第四筆測試紀錄=======================================================

serialNum_4 = driver.find_element(By.XPATH, '//*[@id="content"]/div[4]/div/table/tbody/tr[2]/td')
numText4 = serialNum_4.text
numList4 = list(numText4)
serialNumlist = DataFrame(numList4)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath4 = driver.find_element(By.XPATH, '//*[@id="content"]/div[4]/div/table/tbody/tr[1]/td/div')
textSplit4 = xpath4.text.split()
newList4 = list(textSplit4)
newList4.remove('房間底注:')
newList4.remove('籌碼:')
newList4.remove('總輸贏:')
newDataList4 = DataFrame(newList4)
newList4T = newDataList4.T
csvFile = newList4T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[4]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    # print(textSplit2, '\n')
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[4]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[4]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第五筆測試紀錄=======================================================

serialNum_5 = driver.find_element(By.XPATH, '//*[@id="content"]/div[5]/div/table/tbody/tr[2]/td')
numText5 = serialNum_5.text
numList5 = list(numText5)
serialNumlist = DataFrame(numList5)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath5 = driver.find_element(By.XPATH, '//*[@id="content"]/div[5]/div/table/tbody/tr[1]/td/div')
textSplit5 = xpath5.text.split()
newList5 = list(textSplit5)
newList5.remove('房間底注:')
newList5.remove('籌碼:')
newList5.remove('總輸贏:')
newDataList5 = DataFrame(newList5)
newList5T = newDataList5.T
csvFile = newList5T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[5]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[5]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[5]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第六筆測試紀錄=======================================================

serialNum_6 = driver.find_element(By.XPATH, '//*[@id="content"]/div[6]/div/table/tbody/tr[2]/td')
numText6 = serialNum_6.text
numList6 = list(numText6)
serialNumlist = DataFrame(numList6)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath6 = driver.find_element(By.XPATH, '//*[@id="content"]/div[6]/div/table/tbody/tr[1]/td/div')
textSplit6 = xpath6.text.split()
newList6 = list(textSplit6)
newList6.remove('房間底注:')
newList6.remove('籌碼:')
newList6.remove('總輸贏:')
newDataList6 = DataFrame(newList6)
newList6T = newDataList6.T
csvFile = newList6T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[6]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[6]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[6]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第七筆測試紀錄=======================================================

serialNum_7 = driver.find_element(By.XPATH, '//*[@id="content"]/div[7]/div/table/tbody/tr[2]/td')
numText7 = serialNum_7.text
numList7 = list(numText7)
serialNumlist = DataFrame(numList7)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath7 = driver.find_element(By.XPATH, '//*[@id="content"]/div[7]/div/table/tbody/tr[1]/td/div')
textSplit7 = xpath7.text.split()
newList7 = list(textSplit7)
newList7.remove('房間底注:')
newList7.remove('籌碼:')
newList7.remove('總輸贏:')
newDataList7 = DataFrame(newList7)
newList7T = newDataList7.T
csvFile = newList7T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[7]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[7]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[7]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第八筆測試紀錄=======================================================

serialNum_8 = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div/table/tbody/tr[2]/td')
numText8 = serialNum_8.text
numList8 = list(numText8)
serialNumlist = DataFrame(numList8)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath8 = driver.find_element(By.XPATH, '//*[@id="content"]/div[8]/div/table/tbody/tr[1]/td/div')
textSplit8 = xpath8.text.split()
newList8 = list(textSplit8)
newList8.remove('房間底注:')
newList8.remove('籌碼:')
newList8.remove('總輸贏:')
newDataList8 = DataFrame(newList8)
newList8T = newDataList8.T
csvFile = newList8T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[8]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[8]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[8]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')
      

# =======================================================寫入第九筆測試紀錄=======================================================

serialNum_9 = driver.find_element(By.XPATH, '//*[@id="content"]/div[9]/div/table/tbody/tr[2]/td')
numText9 = serialNum_9.text
numList9 = list(numText9)
serialNumlist = DataFrame(numList9)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath9 = driver.find_element(By.XPATH, '//*[@id="content"]/div[9]/div/table/tbody/tr[1]/td/div')
textSplit9 = xpath9.text.split()
newList9 = list(textSplit9)
newList9.remove('房間底注:')
newList9.remove('籌碼:')
newList9.remove('總輸贏:')
newDataList9 = DataFrame(newList9)
newList9T = newDataList9.T
csvFile = newList9T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[9]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[9]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[9]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')
        

# =======================================================寫入第十筆測試紀錄=======================================================

serialNum_10 = driver.find_element(By.XPATH, '//*[@id="content"]/div[10]/div/table/tbody/tr[2]/td')
numText10 = serialNum_10.text
numList10 = list(numText10)
numList10_1 = ''.join(numList10)
numList10_2 = []
numList10_2.append(numList10_1)
serialNumlist = DataFrame(numList10_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath10 = driver.find_element(By.XPATH, '//*[@id="content"]/div[10]/div/table/tbody/tr[1]/td/div')
textSplit10 = xpath10.text.split()
newList10 = list(textSplit10)
newList10.remove('房間底注:')
newList10.remove('籌碼:')
newList10.remove('總輸贏:')
newDataList10 = DataFrame(newList10)
newList10T = newDataList10.T
csvFile = newList10T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[10]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[10]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[10]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十一筆測試紀錄=======================================================

serialNum_11 = driver.find_element(By.XPATH, '//*[@id="content"]/div[11]/div/table/tbody/tr[2]/td')
numText11 = serialNum_11.text
numList11 = list(numText11)
numList11_1 = ''.join(numList11)
numList11_2 = []
numList11_2.append(numList11_1)
serialNumlist = DataFrame(numList11_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath11 = driver.find_element(By.XPATH, '//*[@id="content"]/div[11]/div/table/tbody/tr[1]/td/div')
textSplit11 = xpath11.text.split()
newList11 = list(textSplit11)
newList11.remove('房間底注:')
newList11.remove('籌碼:')
newList11.remove('總輸贏:')
newDataList11 = DataFrame(newList11)
newList11T = newDataList11.T
csvFile = newList11T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[11]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[11]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[11]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十二筆測試紀錄=======================================================

serialNum_12 = driver.find_element(By.XPATH, '//*[@id="content"]/div[12]/div/table/tbody/tr[2]/td')
numText12 = serialNum_12.text
numList12 = list(numText12)
numList12_1 = ''.join(numList12)
numList12_2 = []
numList12_2.append(numList12_1)
serialNumlist = DataFrame(numList12_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath12 = driver.find_element(By.XPATH, '//*[@id="content"]/div[12]/div/table/tbody/tr[1]/td/div')
textSplit12 = xpath12.text.split()
newList12 = list(textSplit12)
newList12.remove('房間底注:')
newList12.remove('籌碼:')
newList12.remove('總輸贏:')
newDataList12 = DataFrame(newList12)
newList12T = newDataList12.T
csvFile = newList12T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[12]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[12]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[12]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十三筆測試紀錄=======================================================

serialNum_13 = driver.find_element(By.XPATH, '//*[@id="content"]/div[13]/div/table/tbody/tr[2]/td')
numText13 = serialNum_13.text
numList13 = list(numText13)
numList13_1 = ''.join(numList13)
numList13_2 = []
numList13_2.append(numList13_1)
serialNumlist = DataFrame(numList13_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath13 = driver.find_element(By.XPATH, '//*[@id="content"]/div[13]/div/table/tbody/tr[1]/td/div')
textSplit13 = xpath13.text.split()
newList13 = list(textSplit13)
newList13.remove('房間底注:')
newList13.remove('籌碼:')
newList13.remove('總輸贏:')
newDataList13 = DataFrame(newList13)
newList13T = newDataList13.T
csvFile = newList13T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[13]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[13]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[13]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十四筆測試紀錄=======================================================

serialNum_14 = driver.find_element(By.XPATH, '//*[@id="content"]/div[14]/div/table/tbody/tr[2]/td')
numText14 = serialNum_14.text
numList14 = list(numText14)
numList14_1 = ''.join(numList14)
numList14_2 = []
numList14_2.append(numList14_1)
serialNumlist = DataFrame(numList14_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath14 = driver.find_element(By.XPATH, '//*[@id="content"]/div[14]/div/table/tbody/tr[1]/td/div')
textSplit14 = xpath14.text.split()
newList14 = list(textSplit14)
newList14.remove('房間底注:')
newList14.remove('籌碼:')
newList14.remove('總輸贏:')
newDataList14 = DataFrame(newList14)
newList14T = newDataList14.T
csvFile = newList14T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[14]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[14]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[14]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十五筆測試紀錄=======================================================

serialNum_15 = driver.find_element(By.XPATH, '//*[@id="content"]/div[15]/div/table/tbody/tr[2]/td')
numText15 = serialNum_15.text
numList15 = list(numText15)
numList15_1 = ''.join(numList15)
numList15_2 = []
numList15_2.append(numList15_1)
serialNumlist = DataFrame(numList15_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath15 = driver.find_element(By.XPATH, '//*[@id="content"]/div[15]/div/table/tbody/tr[1]/td/div')
textSplit15 = xpath15.text.split()
newList15 = list(textSplit15)
newList15.remove('房間底注:')
newList15.remove('籌碼:')
newList15.remove('總輸贏:')
newDataList15 = DataFrame(newList15)
newList15T = newDataList15.T
csvFile = newList15T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[15]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[15]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[15]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十六筆測試紀錄=======================================================

serialNum_16 = driver.find_element(By.XPATH, '//*[@id="content"]/div[16]/div/table/tbody/tr[2]/td')
numText16 = serialNum_16.text
numList16 = list(numText16)
numList16_1 = ''.join(numList16)
numList16_2 = []
numList16_2.append(numList16_1)
serialNumlist = DataFrame(numList16_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath16 = driver.find_element(By.XPATH, '//*[@id="content"]/div[16]/div/table/tbody/tr[1]/td/div')
textSplit16 = xpath16.text.split()
newList16 = list(textSplit16)
newList16.remove('房間底注:')
newList16.remove('籌碼:')
newList16.remove('總輸贏:')
newDataList16 = DataFrame(newList16)
newList16T = newDataList16.T
csvFile = newList16T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[16]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[16]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[16]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十七筆測試紀錄=======================================================

serialNum_17 = driver.find_element(By.XPATH, '//*[@id="content"]/div[17]/div/table/tbody/tr[2]/td')
numText17 = serialNum_17.text
numList17 = list(numText17)
numList17_1 = ''.join(numList17)
numList17_2 = []
numList17_2.append(numList17_1)
serialNumlist = DataFrame(numList17_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath17 = driver.find_element(By.XPATH, '//*[@id="content"]/div[17]/div/table/tbody/tr[1]/td/div')
textSplit17 = xpath17.text.split()
newList17 = list(textSplit17)
newList17.remove('房間底注:')
newList17.remove('籌碼:')
newList17.remove('總輸贏:')
newDataList17 = DataFrame(newList17)
newList17T = newDataList17.T
csvFile = newList17T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[17]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[17]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[17]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十八筆測試紀錄=======================================================

serialNum_18 = driver.find_element(By.XPATH, '//*[@id="content"]/div[18]/div/table/tbody/tr[2]/td')
numText18 = serialNum_18.text
numList18 = list(numText18)
numList18_1 = ''.join(numList18)
numList18_2 = []
numList18_2.append(numList18_1)
serialNumlist = DataFrame(numList18_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath18 = driver.find_element(By.XPATH, '//*[@id="content"]/div[18]/div/table/tbody/tr[1]/td/div')
textSplit18 = xpath18.text.split()
newList18 = list(textSplit18)
newList18.remove('房間底注:')
newList18.remove('籌碼:')
newList18.remove('總輸贏:')
newDataList18 = DataFrame(newList18)
newList18T = newDataList18.T
csvFile = newList18T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[18]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[18]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[18]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十九筆測試紀錄=======================================================

serialNum_19 = driver.find_element(By.XPATH, '//*[@id="content"]/div[19]/div/table/tbody/tr[2]/td')
numText19 = serialNum_19.text
numList19 = list(numText19)
numList19_1 = ''.join(numList19)
numList19_2 = []
numList19_2.append(numList19_1)
serialNumlist = DataFrame(numList19_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath19 = driver.find_element(By.XPATH, '//*[@id="content"]/div[19]/div/table/tbody/tr[1]/td/div')
textSplit19 = xpath19.text.split()
newList19 = list(textSplit19)
newList19.remove('房間底注:')
newList19.remove('籌碼:')
newList19.remove('總輸贏:')
newDataList19 = DataFrame(newList19)
newList19T = newDataList19.T
csvFile = newList19T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[19]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[19]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[19]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第二十筆測試紀錄=======================================================

serialNum_20 = driver.find_element(By.XPATH, '//*[@id="content"]/div[20]/div/table/tbody/tr[2]/td')
numText20 = serialNum_20.text
numList20 = list(numText20)
numList20_1 = ''.join(numList20)
numList20_2 = []
numList20_2.append(numList20_1)
serialNumlist = DataFrame(numList20_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath20 = driver.find_element(By.XPATH, '//*[@id="content"]/div[20]/div/table/tbody/tr[1]/td/div')
textSplit20 = xpath20.text.split()
newList20 = list(textSplit20)
newList20.remove('房間底注:')
newList20.remove('籌碼:')
newList20.remove('總輸贏:')
newDataList20 = DataFrame(newList20)
newList20T = newDataList20.T
csvFile = newList20T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="content"]/div[20]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f' //*[@id="content"]/div[20]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="content"]/div[20]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')
    
csvRead = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.csv')
excelWrite = csvRead.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.xlsx', header=True, index=True)
print("EXCEL檔案儲存成功!", '\n')

time.sleep(2)


# =======================================================寫入第一筆測試紀錄=======================================================

xpath1 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[6]/td')  # 玩家
xpath2 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[8]/td')  # 局號
xpath3 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[10]/td[1]')  # 結束時間
xpath4 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[10]/td[2]')  # 房間
xpath5 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[12]/td[1]')  # 序號
xpath6 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[12]/td[2]')  # 場景
xpath7 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[14]/td[1]')  # 面額
xpath8 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[14]/td[2]')  # 帳務
xpath9 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[14]/td[3]')  # 押注
xpath10 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[16]/td[1]')  # 彩金
xpath11 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[16]/td[2]')  # 贏分

textSplit1 = xpath1.text.split()
textSplit2 = xpath2.text.split()
textSplit3 = xpath3.text[10:].split()
textSplit4 = xpath4.text.split()
textSplit5 = xpath5.text.split()
textSplit6 = xpath6.text.split()
textSplit7 = xpath7.text.split()
textSplit8 = xpath8.text.split()
textSplit9 = xpath9.text.split()
textSplit10 = xpath10.text.split()
textSplit11 = xpath11.text.split()

newList1 = list(textSplit1)
newList2 = list(textSplit2)
newList3 = list(textSplit3)
newList4 = list(textSplit4)
newList5 = list(textSplit5)
newList6 = list(textSplit6)
newList7 = list(textSplit7)
newList8 = list(textSplit8)
newList9 = list(textSplit9)
newList10 = list(textSplit10)
newList11 = list(textSplit11)
totalList = []
totalList.extend([newList1, newList2, newList3, newList4, newList5, newList6, newList7, newList8, newList9, newList10, newList11])
# print(totalList, '\n')

totalListDT = DataFrame(totalList)
totalListDTT = totalListDT.T

newList1 = DataFrame(newList1)
newList2 = DataFrame(newList2)
newList3 = DataFrame(newList3)
newList4 = DataFrame(newList4)
newList5 = DataFrame(newList5)
newList6 = DataFrame(newList6)
newList7 = DataFrame(newList7)
newList8 = DataFrame(newList8)
newList9 = DataFrame(newList9)
newList10 = DataFrame(newList10)
newList11 = DataFrame(newList11)

csvFileNT = totalListDT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_NT.csv', index=0, encoding='utf-8-sig')
csvFile = totalListDTT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', index=False, encoding='utf-8-sig')

readData = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
newData = readData.rename(columns = {'0':'玩家', '1':'局號', '2':'結束時間', '3':'房間', '4':'序號',
                                     '5':'場景', '6':'面額', '7':'帳務', '8':'押注', '9':'彩金', '10':'贏分'})
newFile1 = newData.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', index=False, encoding = 'utf-8-sig')

excelFileNT = totalListDT.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_NT.xlsx', index=None, header=True)
# newList1T = newList1.T
# csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).csv', index=0, encoding='utf-8-sig')
print('第', 1, '筆資料寫入完成。', '\n')


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[6]/td'  # 玩家姓名
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False) 


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[8]/td'  # 局號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[10]/td[1]'  # 結束時間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text[10:].split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[10]/td[2]'  # 房間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[12]/td[1]'  # 序號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[12]/td[2]'  # 場景
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)
   

for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[14]/td[1]'  # 面額
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[14]/td[2]'  # 帳務
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[14]/td[3]'  # 押注
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[16]/td[1]'  # 彩金
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)


for i in range(2, 21):
    xpath = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[16]/td[2]'  # 贏分 
    element = driver.find_element(By.XPATH, xpath) 
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.csv')
    # readEX = pd.read_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913).xlsx')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)
    print('第', i, '筆資料寫入完成。', '\n')

time.sleep(2)
driver.close()

workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx')
source_sheet = workbook['Sheet1']

source_sheet.move_range('A22' ":" 'A40', rows = -19, cols = 1)
source_sheet.move_range('A41' ":" 'A59', rows = -38, cols = 2)
source_sheet.move_range('A60' ":" 'A78', rows = -57, cols = 3)
source_sheet.move_range('A79' ":" 'A97', rows = -76, cols = 4)
source_sheet.move_range('A98' ":" 'A116', rows = -95, cols = 5)
source_sheet.move_range('A117' ":" 'A135', rows = -114, cols = 6)
source_sheet.move_range('A136' ":" 'A154', rows = -133, cols = 7)
source_sheet.move_range('A155' ":" 'A173', rows = -152, cols = 8)
source_sheet.move_range('A174' ":" 'A192', rows = -171, cols = 9)
source_sheet.move_range('A193' ":" 'A211', rows = -190, cols = 10)
source_sheet.move_range('A212' ":" 'A230', rows = -209, cols = 11)
source_sheet.move_range('A231' ":" 'A249', rows = -228, cols = 12)
source_sheet.move_range('A250' ":" 'A268', rows = -247, cols = 13)
 
workbook.save(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx')
workbook.close()

print("EXCEL檔案儲存成功!", '\n')
time.sleep(2)   

print('(菲律賓骰寶)前台共20筆遊戲紀錄獲取完成!', '\n')


# ============================Section.1 進入後台首頁============================

# ch_options = Options()
# ch_options.add_argument("--headless")  #無WEB UI顯示
# driver = webdriver.Chrome(ch_options)

driver = webdriver.Chrome()  # 有WEB UI顯示

time.sleep(1) 
driver.get("https://dev-admin-br-02.claretfox.com/")
http_status = requests.get('https://dev-admin-br-02.claretfox.com/')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print('Http Response Code:', stateCode, '\n')
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')
time.sleep(2)

logging.basicConfig(level=logging.DEBUG,
                    filename='output.log',
                    datefmt='%Y/%m/%d %H:%M:%S',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(lineno)d - %(module)s - %(message)s')
logger = logging.getLogger(__name__)

logger.info('This is a log info')
logger.debug('Debugging')
logger.warning('Warning exists')
logger.info('Finish')



back_platform = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div/div/div/form/div/div[1]")))
back_platform.click()
time.sleep(2)
print("進入DEV後台首頁!", '\n')

# -----------切換網頁顯示語系-----------
language_bar = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/form/label')))
language_bar.click()

language_ch_zh = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="lang"]/option[2]')))
language_ch_zh.click()
print("語系已切換'繁體中文'!", '\n')

# ============================Section.2 登入使用者名稱及密碼============================

login_icon = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, 'content-group')))
login_icon.click()  # 尋找登入介面元素位址
time.sleep(2)

login_acc = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="user_id"]')))
login_acc.clear()    # 預設此欄位為null, 但仍先清除帳號欄位資訊
login_pass = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]')))   
login_pass.clear()   # 預設此欄位為null, 但仍先清除密碼欄位資訊

login_verificationCode = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input')))  # 驗證碼欄位

verificationCode = driver.find_element(By.CSS_SELECTOR, '#captcha_img > img')
verificationCode.screenshot('verificationCode.png')
time.sleep(1)

actions = ActionChains(driver)
actions.move_to_element(login_acc)    
actions.perform()
login_acc.send_keys("ivan_li")   # 個人使用者帳號
time.sleep(1)
login_acc.send_keys(Keys.TAB)   # 切換至密碼輸入欄位
login_pass.send_keys("iPlaystar296")  # 個人密碼
time.sleep(1)
login_pass.send_keys(Keys.TAB)   # 切換至驗證碼輸入欄位
ocr = ddddocr.DdddOcr()
with open('verificationCode.png', 'rb') as fp:
    image = fp.read()
catch = ocr.classification(image)
login_verificationCode.send_keys(catch)
time.sleep(1)

login_button = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[6]/button')))
login_button.click()
print("登入成功!", '\n')
time.sleep(2)

# ============================Section.3 切換後台功能頁籤============================

player = driver.find_element(By.ID, 'Player')
player.click()
print("進入玩家功能選單!", '\n')
time.sleep(2)

driver.get('https://dev-admin-br-02.claretfox.com/Player/game_history')
# /html/body/div[3]/div/div[2]/div[1]/div/div/ul/li[3]/ul/li[3]/a
# actions.move_to_element(playerRecord)    
# actions.perform()    
# playerRecord.click()
print("切換遊戲紀錄子選單!", '\n')
time.sleep(2)

start_time = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div')))

startTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button')
actions.move_to_element(startTime_button)    
actions.perform()
startTime_button.click()

startDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(startDate)    
actions.perform()
startDate.click()

startTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[5]/a')  # 開始日期選取 "2024.09.13"
startTime_check.click()
time.sleep(1)

startTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
startTime_confirm.click()
time.sleep(2)

endTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i')
actions.move_to_element(endTime_button)    
actions.perform()
endTime_button.click()

endDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
actions.move_to_element(endDate)    
actions.perform()
endDate.click()

endTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[5]/a')  # 結束日期選取 "2024.09.13"
endTime_check.click()
time.sleep(1)

endTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
endTime_confirm.click()
time.sleep(2)

showLimit = driver.find_element(By.XPATH, '//*[@id="count"]')
showLimit.click()
time.sleep(2)

showLimitSet = driver.find_element(By.XPATH, '//*[@id="count"]/option[2]')  # 設定顯示筆數 = 50筆
showLimitSet.click()
time.sleep(2)

agencyType = driver.find_element(By.XPATH, '//*[@id="agent_attr"]')
agencyType.click()
time.sleep(2)

agencyTypeSet = driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[10]')  # 設定代理商類別 = Test
agencyTypeSet.click()
time.sleep(2)
print('代理商類別: ', agencyTypeSet.text, '\n')

agencyName = driver.find_element(By.XPATH, '//*[@id="agent"]')
agencyName.click()
time.sleep(2)

agencyNameSet = driver.find_element(By.XPATH, '//*[@id="agent"]/option[7]')  # 設定代理商名稱 = Test-2
agencyNameSet.click()
time.sleep(2)
print('代理商名稱: ', agencyNameSet.text, '\n')

gameType = driver.find_element(By.XPATH, '//*[@id="game_type"]')
gameType.click()
time.sleep(2)

gameTypeSet = driver.find_element(By.XPATH, '//*[@id="game_type"]/option[7]')  # 設定遊戲類別 = 棋牌遊戲
driver.execute_script("arguments[0].scrollIntoView()", gameTypeSet)
gameTypeSet.click()
time.sleep(2)
print('遊戲類別: ', gameTypeSet.text, '\n')

gameSelect = driver.find_element(By.XPATH, '//*[@id="game_select"]')
gameSelect.click()
time.sleep(2)

gameSelectSet = driver.find_element(By.XPATH, '//*[@id="game_select"]/option[8]')  # 遊戲選擇 = 菲律賓骰寶
gameSelectSet.click()
time.sleep(2)
print('遊戲名稱: ', gameSelectSet.text, '\n')

playerName = driver.find_element(By.XPATH, '//*[@id="player"]')  # 玩家名稱 = 'ivan_li'
playerName.send_keys("ivan_li")
time.sleep(2)

btnSubmit = driver.find_element(By.XPATH, '//*[@id="sh_btn"]')  # 確認無誤送出
btnSubmit.click()
time.sleep(10)

print('獲取(菲律賓骰寶)後台遊戲紀錄中...', '\n')

# =======================================================寫入第一筆測試紀錄=======================================================

serialNum_1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[2]/td/div')
numText1 = serialNum_1.text
numList1 = list(numText1)
serialNumlist = DataFrame(numList1)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏'])

xpath1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[1]/td/div')  # 房間底注 / 籌碼 / 總輸贏 元素位置
textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newList1.remove('房間底注:')
newList1.remove('籌碼:')
newList1.remove('總輸贏:')     
            
newDataList1 = DataFrame(newList1)
newList1T = newDataList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')
    
for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[1]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    # print(textSplit2, '\n')
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile2 = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[1]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[1]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第二筆測試紀錄=======================================================

serialNum_2 = driver.find_element(By.XPATH, '//*[@id="history"]/div[2]/div/table/tbody/tr[2]/td/div')
numText2 = serialNum_2.text
numList2 = list(numText2)
serialNumlist = DataFrame(numList2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏'])
 
xpath2 = driver.find_element(By.XPATH, '//*[@id="history"]/div[2]/div/table/tbody/tr[1]/td/div')
textSplit2 = xpath2.text.split()
newList2 = list(textSplit2)
# print(newList1, '\n')
newList2.remove('房間底注:')
newList2.remove('籌碼:')
newList2.remove('總輸贏:')
# print(newList1, '\n')
newDataList2 = DataFrame(newList2)
newList2T = newDataList2.T
csvFile = newList2T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[2]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    # print(textSplit2, '\n')
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile2 = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[2]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[2]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第三筆測試紀錄=======================================================

serialNum_3 = driver.find_element(By.XPATH, '//*[@id="history"]/div[3]/div/table/tbody/tr[2]/td/div')
numText3 = serialNum_3.text
numList3 = list(numText3)
serialNumlist = DataFrame(numList3)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath3 = driver.find_element(By.XPATH, '//*[@id="history"]/div[3]/div/table/tbody/tr[1]/td/div')
textSplit3 = xpath3.text.split()
newList3 = list(textSplit3)
# print(newList1, '\n')
newList3.remove('房間底注:')
newList3.remove('籌碼:')
newList3.remove('總輸贏:')
# print(newList1, '\n')
newDataList3 = DataFrame(newList3)
newList3T = newDataList3.T
csvFile = newList3T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[3]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    # print(textSplit2, '\n')
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile2 = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[3]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[3]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第四筆測試紀錄=======================================================

serialNum_4 = driver.find_element(By.XPATH, '//*[@id="history"]/div[4]/div/table/tbody/tr[2]/td/div')
numText4 = serialNum_4.text
numList4 = list(numText4)
serialNumlist = DataFrame(numList4)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath4 = driver.find_element(By.XPATH, '//*[@id="history"]/div[4]/div/table/tbody/tr[1]/td/div')
textSplit4 = xpath4.text.split()
newList4 = list(textSplit4)
newList4.remove('房間底注:')
newList4.remove('籌碼:')
newList4.remove('總輸贏:')
newDataList4 = DataFrame(newList4)
newList4T = newDataList4.T
csvFile = newList4T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[4]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    # print(textSplit2, '\n')
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[4]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[4]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第五筆測試紀錄=======================================================

serialNum_5 = driver.find_element(By.XPATH, '//*[@id="history"]/div[5]/div/table/tbody/tr[2]/td/div')
numText5 = serialNum_5.text
numList5 = list(numText5)
serialNumlist = DataFrame(numList5)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath5 = driver.find_element(By.XPATH, '//*[@id="history"]/div[5]/div/table/tbody/tr[1]/td/div')
textSplit5 = xpath5.text.split()
newList5 = list(textSplit5)
newList5.remove('房間底注:')
newList5.remove('籌碼:')
newList5.remove('總輸贏:')
newDataList5 = DataFrame(newList5)
newList5T = newDataList5.T
csvFile = newList5T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[5]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[5]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[5]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第六筆測試紀錄=======================================================

serialNum_6 = driver.find_element(By.XPATH, '//*[@id="history"]/div[6]/div/table/tbody/tr[2]/td/div')
numText6 = serialNum_6.text
numList6 = list(numText6)
serialNumlist = DataFrame(numList6)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath6 = driver.find_element(By.XPATH, '//*[@id="history"]/div[6]/div/table/tbody/tr[1]/td/div')
textSplit6 = xpath6.text.split()
newList6 = list(textSplit6)
newList6.remove('房間底注:')
newList6.remove('籌碼:')
newList6.remove('總輸贏:')
newDataList6 = DataFrame(newList6)
newList6T = newDataList6.T
csvFile = newList6T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[6]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[6]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[6]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第七筆測試紀錄=======================================================

serialNum_7 = driver.find_element(By.XPATH, '//*[@id="history"]/div[7]/div/table/tbody/tr[2]/td/div')
numText7 = serialNum_7.text
numList7 = list(numText7)
serialNumlist = DataFrame(numList7)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath7 = driver.find_element(By.XPATH, '//*[@id="history"]/div[7]/div/table/tbody/tr[1]/td/div')
textSplit7 = xpath7.text.split()
newList7 = list(textSplit7)
newList7.remove('房間底注:')
newList7.remove('籌碼:')
newList7.remove('總輸贏:')
newDataList7 = DataFrame(newList7)
newList7T = newDataList7.T
csvFile = newList7T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[7]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[7]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[7]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第八筆測試紀錄=======================================================

serialNum_8 = driver.find_element(By.XPATH, '//*[@id="history"]/div[8]/div/table/tbody/tr[2]/td/div')
numText8 = serialNum_8.text
numList8 = list(numText8)
serialNumlist = DataFrame(numList8)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath8 = driver.find_element(By.XPATH, '//*[@id="history"]/div[8]/div/table/tbody/tr[1]/td/div')
textSplit8 = xpath8.text.split()
newList8 = list(textSplit8)
newList8.remove('房間底注:')
newList8.remove('籌碼:')
newList8.remove('總輸贏:')
newDataList8 = DataFrame(newList8)
newList8T = newDataList8.T
csvFile = newList8T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[8]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[8]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[8]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第九筆測試紀錄=======================================================

serialNum_9 = driver.find_element(By.XPATH, '//*[@id="history"]/div[9]/div/table/tbody/tr[2]/td/div')
numText9 = serialNum_9.text
numList9 = list(numText9)
serialNumlist = DataFrame(numList9)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath9 = driver.find_element(By.XPATH, '//*[@id="history"]/div[9]/div/table/tbody/tr[1]/td/div')
textSplit9 = xpath9.text.split()
newList9 = list(textSplit9)
newList9.remove('房間底注:')
newList9.remove('籌碼:')
newList9.remove('總輸贏:')
newDataList9 = DataFrame(newList9)
newList9T = newDataList9.T
csvFile = newList9T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[9]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[9]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[9]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十筆測試紀錄=======================================================

serialNum_10 = driver.find_element(By.XPATH, '//*[@id="history"]/div[10]/div/table/tbody/tr[2]/td/div')
numText10 = serialNum_10.text
numList10 = list(numText10)
numList10_1 = ''.join(numList10)
numList10_2 = []
numList10_2.append(numList10_1)
serialNumlist = DataFrame(numList10_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath10 = driver.find_element(By.XPATH, '//*[@id="history"]/div[10]/div/table/tbody/tr[1]/td/div')
textSplit10 = xpath10.text.split()
newList10 = list(textSplit10)
newList10.remove('房間底注:')
newList10.remove('籌碼:')
newList10.remove('總輸贏:')
newDataList10 = DataFrame(newList10)
newList10T = newDataList10.T
csvFile = newList10T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[10]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[10]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[10]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十一筆測試紀錄=======================================================

serialNum_11 = driver.find_element(By.XPATH, '//*[@id="history"]/div[11]/div/table/tbody/tr[2]/td/div')
numText11 = serialNum_11.text
numList11 = list(numText11)
numList11_1 = ''.join(numList11)
numList11_2 = []
numList11_2.append(numList11_1)
serialNumlist = DataFrame(numList11_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath11 = driver.find_element(By.XPATH, '//*[@id="history"]/div[11]/div/table/tbody/tr[1]/td/div')
textSplit11 = xpath11.text.split()
newList11 = list(textSplit11)
newList11.remove('房間底注:')
newList11.remove('籌碼:')
newList11.remove('總輸贏:')
newDataList11 = DataFrame(newList11)
newList11T = newDataList11.T
csvFile = newList11T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[11]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[11]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[11]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十二筆測試紀錄=======================================================

serialNum_12 = driver.find_element(By.XPATH, '//*[@id="history"]/div[12]/div/table/tbody/tr[2]/td/div')
numText12 = serialNum_12.text
numList12 = list(numText12)
numList12_1 = ''.join(numList12)
numList12_2 = []
numList12_2.append(numList12_1)
serialNumlist = DataFrame(numList12_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath12 = driver.find_element(By.XPATH, '//*[@id="history"]/div[12]/div/table/tbody/tr[1]/td/div')
textSplit12 = xpath12.text.split()
newList12 = list(textSplit12)
newList12.remove('房間底注:')
newList12.remove('籌碼:')
newList12.remove('總輸贏:')
newDataList12 = DataFrame(newList12)
newList12T = newDataList12.T
csvFile = newList12T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[12]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[12]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[12]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十三筆測試紀錄=======================================================

serialNum_13 = driver.find_element(By.XPATH, '//*[@id="history"]/div[13]/div/table/tbody/tr[2]/td/div')
numText13 = serialNum_13.text
numList13 = list(numText13)
numList13_1 = ''.join(numList13)
numList13_2 = []
numList13_2.append(numList13_1)
serialNumlist = DataFrame(numList13_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath13 = driver.find_element(By.XPATH, '//*[@id="history"]/div[13]/div/table/tbody/tr[1]/td/div')
textSplit13 = xpath13.text.split()
newList13 = list(textSplit13)
newList13.remove('房間底注:')
newList13.remove('籌碼:')
newList13.remove('總輸贏:')
newDataList13 = DataFrame(newList13)
newList13T = newDataList13.T
csvFile = newList13T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[13]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[13]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[13]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十四筆測試紀錄=======================================================

serialNum_14 = driver.find_element(By.XPATH, '//*[@id="history"]/div[14]/div/table/tbody/tr[2]/td/div')
numText14 = serialNum_14.text
numList14 = list(numText14)
numList14_1 = ''.join(numList14)
numList14_2 = []
numList14_2.append(numList14_1)
serialNumlist = DataFrame(numList14_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath14 = driver.find_element(By.XPATH, '//*[@id="history"]/div[14]/div/table/tbody/tr[1]/td/div')
textSplit14 = xpath14.text.split()
newList14 = list(textSplit14)
newList14.remove('房間底注:')
newList14.remove('籌碼:')
newList14.remove('總輸贏:')
newDataList14 = DataFrame(newList14)
newList14T = newDataList14.T
csvFile = newList14T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[14]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[14]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[14]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十五筆測試紀錄=======================================================

serialNum_15 = driver.find_element(By.XPATH, '//*[@id="history"]/div[15]/div/table/tbody/tr[2]/td/div')
numText15 = serialNum_15.text
numList15 = list(numText15)
numList15_1 = ''.join(numList15)
numList15_2 = []
numList15_2.append(numList15_1)
serialNumlist = DataFrame(numList15_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath15 = driver.find_element(By.XPATH, '//*[@id="history"]/div[15]/div/table/tbody/tr[1]/td/div')
textSplit15 = xpath15.text.split()
newList15 = list(textSplit15)
newList15.remove('房間底注:')
newList15.remove('籌碼:')
newList15.remove('總輸贏:')
newDataList15 = DataFrame(newList15)
newList15T = newDataList15.T
csvFile = newList15T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[15]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[15]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[15]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十六筆測試紀錄=======================================================

serialNum_16 = driver.find_element(By.XPATH, '//*[@id="history"]/div[16]/div/table/tbody/tr[2]/td/div')
numText16 = serialNum_16.text
numList16 = list(numText16)
numList16_1 = ''.join(numList16)
numList16_2 = []
numList16_2.append(numList16_1)
serialNumlist = DataFrame(numList16_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath16 = driver.find_element(By.XPATH, '//*[@id="history"]/div[16]/div/table/tbody/tr[1]/td/div')
textSplit16 = xpath16.text.split()
newList16 = list(textSplit16)
newList16.remove('房間底注:')
newList16.remove('籌碼:')
newList16.remove('總輸贏:')
newDataList16 = DataFrame(newList16)
newList16T = newDataList16.T
csvFile = newList16T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[16]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[16]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[16]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十七筆測試紀錄=======================================================

serialNum_17 = driver.find_element(By.XPATH, '//*[@id="history"]/div[17]/div/table/tbody/tr[2]/td/div')
numText17 = serialNum_17.text
numList17 = list(numText17)
numList17_1 = ''.join(numList17)
numList17_2 = []
numList17_2.append(numList17_1)
serialNumlist = DataFrame(numList17_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath17 = driver.find_element(By.XPATH, '//*[@id="history"]/div[17]/div/table/tbody/tr[1]/td/div')
textSplit17 = xpath17.text.split()
newList17 = list(textSplit17)
newList17.remove('房間底注:')
newList17.remove('籌碼:')
newList17.remove('總輸贏:')
newDataList17 = DataFrame(newList17)
newList17T = newDataList17.T
csvFile = newList17T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[17]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[17]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[17]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十八筆測試紀錄=======================================================

serialNum_18 = driver.find_element(By.XPATH, '//*[@id="history"]/div[18]/div/table/tbody/tr[2]/td/div')
numText18 = serialNum_18.text
numList18 = list(numText18)
numList18_1 = ''.join(numList18)
numList18_2 = []
numList18_2.append(numList18_1)
serialNumlist = DataFrame(numList18_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath18 = driver.find_element(By.XPATH, '//*[@id="history"]/div[18]/div/table/tbody/tr[1]/td/div')
textSplit18 = xpath18.text.split()
newList18 = list(textSplit18)
newList18.remove('房間底注:')
newList18.remove('籌碼:')
newList18.remove('總輸贏:')
newDataList18 = DataFrame(newList18)
newList18T = newDataList18.T
csvFile = newList18T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[18]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[18]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[18]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第十九筆測試紀錄=======================================================

serialNum_19 = driver.find_element(By.XPATH, '//*[@id="history"]/div[19]/div/table/tbody/tr[2]/td/div')
numText19 = serialNum_19.text
numList19 = list(numText19)
numList19_1 = ''.join(numList19)
numList19_2 = []
numList19_2.append(numList19_1)
serialNumlist = DataFrame(numList19_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath19 = driver.find_element(By.XPATH, '//*[@id="history"]/div[19]/div/table/tbody/tr[1]/td/div')
textSplit19 = xpath19.text.split()
newList19 = list(textSplit19)
newList19.remove('房間底注:')
newList19.remove('籌碼:')
newList19.remove('總輸贏:')
newDataList19 = DataFrame(newList19)
newList19T = newDataList19.T
csvFile = newList19T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[19]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[19]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[19]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


# =======================================================寫入第二十筆測試紀錄=======================================================

serialNum_20 = driver.find_element(By.XPATH, '//*[@id="history"]/div[20]/div/table/tbody/tr[2]/td/div')
numText20 = serialNum_20.text
numList20 = list(numText20)
numList20_1 = ''.join(numList20)
numList20_2 = []
numList20_2.append(numList20_1)
serialNumlist = DataFrame(numList20_2)
csvFile = serialNumlist.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

with open(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(['房間底注', '籌碼', '總輸贏']) 
 
xpath20 = driver.find_element(By.XPATH, '//*[@id="history"]/div[20]/div/table/tbody/tr[1]/td/div')
textSplit20 = xpath20.text.split()
newList20 = list(textSplit20)
newList20.remove('房間底注:')
newList20.remove('籌碼:')
newList20.remove('總輸贏:')
newDataList20 = DataFrame(newList20)
newList20T = newDataList20.T
csvFile = newList20T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=False, encoding='utf-8-sig')

for i in range(1, 10):
    xpath_ML = f'//*[@id="history"]/div[20]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[1]/div[{i}]'  # 中間區塊左側 '押注'/'贏分' 紀錄
    element_ML = driver.find_element(By.XPATH, xpath_ML)
    textSplit_ML = element_ML.text.split()
    
    newList_ML = list(textSplit_ML)
    newDataList = []
    for x in newList_ML:
        if 'x' not in x:
            newDataList.append(x)
        if x == '3' or x == 'Same' or x == 'Color':
            newDataList.remove(x)
            
    newDataList_ML = DataFrame(newDataList)
    newList_MLT = newDataList_ML.T
    csvFile = newList_MLT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', header=False, index=None, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MM = f'//*[@id="history"]/div[20]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[2]/div[{i}]'  # 中間區塊中間側 '押注'/'贏分' 紀錄
    element_MM = driver.find_element(By.XPATH, xpath_MM)
    textSplit_MM = element_MM.text.split()
    
    newList_MM = list(textSplit_MM)
    newDataList = []
    for x in newList_MM:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MM = DataFrame(newDataList)
    newList_MMT = newDataList_MM.T
    csvFile = newList_MMT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')


for i in range(1, 10):
    xpath_MR = f'//*[@id="history"]/div[20]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]/div[3]/div[{i}]'  # 中間區塊右側 '押注'/'贏分' 紀錄
    element_MR = driver.find_element(By.XPATH, xpath_MR)
    textSplit_MR = element_MR.text.split()
    
    newList_MR = list(textSplit_MR)
    newDataList = []
    for x in newList_MR:
        if 'x' not in x:
            newDataList.append(x)
    
    newDataList_MR = DataFrame(newDataList)
    newList_MRT = newDataList_MR.T
    csvFile = newList_MRT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv', mode='a', index=False, header=False, encoding='utf-8-sig')
    
csvRead = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.csv')
excelWrite = csvRead.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.xlsx', header=True, index=True)
print("EXCEL檔案儲存成功!", '\n')

time.sleep(2)

# =======================================================寫入第一筆測試紀錄=======================================================

xpath1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[6]/td')  # 玩家
xpath2 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[8]/td')  # 局號
xpath3 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[10]/td[2]')  # 結束時間
xpath4 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[12]/td[3]')  # 房間
xpath5 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[14]/td[1]')  # 序號
xpath6 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[12]/td[2]')  # 場景
xpath7 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[14]/td[2]')  # 面額
xpath8 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[14]/td[3]')  # 帳務
xpath9 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[14]/td[4]')  # 押注
xpath10 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[16]/td[1]')  # 彩金
xpath11 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[16]/td[2]')  # 贏分

textSplit1 = xpath1.text[8:].split()
textSplit2 = xpath2.text.split()
textSplit3 = xpath3.text[10:].split()
textSplit4 = xpath4.text.split()
textSplit5 = xpath5.text.split()
textSplit6 = xpath6.text.split()
textSplit7 = xpath7.text.split()
textSplit8 = xpath8.text.split()
textSplit9 = xpath9.text.split()
textSplit10 = xpath10.text.split()
textSplit11 = xpath11.text.split()

newList1 = list(textSplit1)
newList2 = list(textSplit2)
newList3 = list(textSplit3)
newList4 = list(textSplit4)
newList5 = list(textSplit5)
newList6 = list(textSplit6)
newList7 = list(textSplit7)
newList8 = list(textSplit8)
newList9 = list(textSplit9)
newList10 = list(textSplit10)
newList11 = list(textSplit11)

totalList = []
totalList.extend([newList1, newList2, newList3, newList4, newList5, newList6, newList7, newList8, newList9, newList10, newList11])

totalListDT = DataFrame(totalList)
totalListDTT = totalListDT.T

newList1 = DataFrame(newList1)
newList2 = DataFrame(newList2)
newList3 = DataFrame(newList3)
newList4 = DataFrame(newList4)
newList5 = DataFrame(newList5)
newList6 = DataFrame(newList6)
newList7 = DataFrame(newList7)
newList8 = DataFrame(newList8)
newList9 = DataFrame(newList9)
newList10 = DataFrame(newList10)
newList11 = DataFrame(newList11)

csvFileNT = totalListDT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_NT.csv', index=0, encoding='utf-8-sig')
csvFile = totalListDTT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', index=False, encoding='utf-8-sig')

readData = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
newData = readData.rename(columns = {'0':'玩家', '1':'局號', '2':'結束時間', '3':'房間', '4':'序號',
                                     '5':'場景', '6':'面額', '7':'帳務', '8':'押注', '9':'彩金', '10':'贏分'})
newFile1 = newData.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', index=False, encoding = 'utf-8-sig')
excelFileNT = totalListDT.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_NT.xlsx', index=None, header=True)

print('第', 1, '筆資料寫入完成。', '\n')


for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[6]/td'  # 玩家姓名
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text[8:].split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False) 
    
for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[8]/td'  # 局號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)
    
for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[10]/td[2]'  # 結束時間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text[10:].split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)
    
for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[12]/td[3]'  # 房間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)
    
for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[14]/td[1]'  # 序號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)

for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[12]/td[2]'  # 場景
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)
   
for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[14]/td[2]'  # 面額
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)

for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[14]/td[3]'  # 帳務
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)
    
for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[14]/td[4]'  # 押注
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)

for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[16]/td[1]'  # 彩金
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)

for i in range(2, 21):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[16]/td[2]'  # 贏分 
    element = driver.find_element(By.XPATH, xpath) 
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx', header=True, index=False)
    print('第', i, '筆資料寫入完成。', '\n')
    
print("EXCEL檔案儲存成功!", '\n')
print('(菲律賓骰寶)後台共20筆遊戲紀錄獲取完成!', '\n')
time.sleep(2)   
driver.close()


workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx')
source_sheet = workbook['Sheet1']

source_sheet.move_range('A22' ":" 'A40', rows = -19, cols = 1)
source_sheet.move_range('A41' ":" 'A59', rows = -38, cols = 2)
source_sheet.move_range('A60' ":" 'A78', rows = -57, cols = 3)
source_sheet.move_range('A79' ":" 'A97', rows = -76, cols = 4)
source_sheet.move_range('A98' ":" 'A116', rows = -95, cols = 5)
source_sheet.move_range('A117' ":" 'A135', rows = -114, cols = 6)
source_sheet.move_range('A136' ":" 'A154', rows = -133, cols = 7)
source_sheet.move_range('A155' ":" 'A173', rows = -152, cols = 8)
source_sheet.move_range('A174' ":" 'A192', rows = -171, cols = 9)
source_sheet.move_range('A193' ":" 'A211', rows = -190, cols = 10)

workbook.save(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx')
workbook.close()


# =================================================== Step.3 前/後台資料分析 ===================================================

print('(菲律賓骰寶)前/後台遊戲紀錄比對中...', '\n')

# 1.數據比對

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Game_20240913)_upper.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Game_20240913)_upper.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20240913)_upper_analysis.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20240913)_upper_analysis.xlsx')
workbook_2.close()
time.sleep(1)
print('(菲律賓骰寶)前/後台遊戲紀錄比對完成!', '\n')


# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20240913)_upper_analysis.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20240913)_upper_analysis.xlsx')

print('(菲律賓骰寶)前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_upper_20240913.xlsx")
print("(菲律賓骰寶)前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = "C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_upper_20240913.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '前台遊戲紀錄比對結果(菲律賓骰寶)_20240913'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '後台遊戲紀錄比對結果(菲律賓骰寶)_20240913'  # 修改分頁.2工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')

# 3.工作表顏色設定

wb = openpyxl.load_workbook(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_upper_20240913.xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
wb.save(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_upper_20240913.xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()


# =================================================== Step.3 前/後台資料分析 ===================================================

print('(菲律賓骰寶)前/後台遊戲紀錄比對中...', '\n')

# 1.數據比對

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20240913)_lower.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20240913)_lower.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20240913)_lower_analysis.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20240913)_lower_analysis.xlsx')
workbook_2.close()
time.sleep(1)
print('(菲律賓骰寶)前/後台遊戲紀錄比對完成!', '\n')


# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20240913)_lower_analysis.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20240913)_lower_analysis.xlsx')

print('(菲律賓骰寶)前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_lower_20240913.xlsx")
print("(菲律賓骰寶)前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = "C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_lower_20240913.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '前台遊戲紀錄比對結果(菲律賓骰寶)_20240913'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '後台遊戲紀錄比對結果(菲律賓骰寶)_20240913'  # 修改分頁.2工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')

# 3.工作表顏色設定

wb = openpyxl.load_workbook(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_lower_20240913.xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
wb.save(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_lower_20240913.xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()

print('(菲律賓骰寶)前/後台共20筆遊戲紀錄擷取(含數據比對)測試完成!', '\n')

logging.basicConfig(level=logging.DEBUG,
                    filename='output.log',
                    datefmt='%Y/%m/%d %H:%M:%S',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(lineno)d - %(module)s - %(message)s')
logger = logging.getLogger(__name__)

logger.info('This is a log info')
logger.debug('Debugging')
logger.warning('Warning exists')
logger.info('Finish')

end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
print('測試結束時間: ', end_time, '\n')

time.sleep(2)
driver.quit()