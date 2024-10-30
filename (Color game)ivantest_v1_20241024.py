# 套件匯入
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pyautogui as pag
from selenium.webdriver.chrome.service import Service
import requests
from http import HTTPStatus
from selenium.webdriver.chrome.options import Options
from lxml import html
import csv
import numpy as np
import pandas as pd
import cv2
import pybi as pbi
import os
import sys
import aspose
import xlrd
import xlwt
import glob
import cx_Freeze
from cx_Freeze import setup
import setuptools
import jpype
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType
import difflib
import openpyxl
from openpyxl import Workbook
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl.styles import Font  # 導入字體模組
from openpyxl.styles import PatternFill  # 導入填充模組
from spire.xls import *
from collections import deque
from pandas.core.frame import DataFrame
# 导入openpyxl模块并将其重命名为pxl
import openpyxl as pxl
# 从openpyxl导入PatternFill类
from openpyxl.styles import PatternFill
# 从openpyxl导入colors类
from openpyxl.styles import colors
# 从openpyxl导入Font类
from openpyxl.styles import Font
import datetime
import pytesseract
from PIL import Image
import ddddocr
import logging
import re


current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
print('測試起始時間: ', current_time, '\n')

print('串接(菲律賓骰寶)前台API擷取遊戲紀錄測試中...', '\n')
time.sleep(1)

# ch_options = Options()
# ch_options.add_argument("--headless")  # 無WEB UI顯示
# driver = webdriver.Chrome(ch_options)
   
driver = webdriver.Chrome() 
driver.get('https://dev-api.iplaystar.net/gamehistory/?host_id=aa62ffb88b300f6be6654615f17ce6fa&lang=tch&game_id=PSC-ON-00016&count=20&page=1&uid=uN22JyloR1aD2zgUIu8nL1ogWFBPqrujwHHpugHyw94%3d')
http_status = requests.get('https://dev-api.iplaystar.net/gamehistory/?host_id=aa62ffb88b300f6be6654615f17ce6fa&lang=tch&game_id=PSC-ON-00016&count=20&page=1&uid=uN22JyloR1aD2zgUIu8nL1ogWFBPqrujwHHpugHyw94%3d')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')
time.sleep(2)

# =================================選取記錄日期 =================================

dateList = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="serdate"]')))  # 點選日期選單
actions = ActionChains(driver)
actions.move_to_element(dateList)    
actions.perform()
dateList.click()
time.sleep(2)

# recordDate = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/thead/tr[2]/th[1]')  # 選取 "<"鍵跳至前月日期選單
# actions.move_to_element(recordDate)    
# actions.perform()
# recordDate.click()

date_select01 = driver.find_element(By.XPATH, '/html/body/div[7]/div[1]/table/tbody/tr[4]/td[5]')  # 點選日期(2024.10.24)
date_select01.click()
time.sleep(1)
print('獲取(菲律賓骰寶)前台遊戲紀錄中...', '\n')

path_1 = r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data'
if not os.path.isdir(path_1):
    os.makedirs(path_1)
    
# =======================================================寫入第1筆測試紀錄(房間底注/籌碼/總輸贏)=======================================================

# serialNum_1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[2]/td/div')
# numText1 = serialNum_1.text
# numList1 = list(numText1)
# serialNumlist = DataFrame(numList1)
# csvFile = serialNumlist.to_csv(r'C:\Users\ivan_li\Desktop\(3). Color Game\(1). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper.csv', header=False, index=False, encoding='utf-8-sig')

# with open(r'C:\Users\ivan_li\Desktop\(3). Color Game\(1). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper.csv', mode='a', newline='', encoding='utf-8-sig') as csvfile:
#     writer = csv.writer(csvfile)
#     writer.writerow(['房間底注', '籌碼', '總輸贏'])

xpath1 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[1]/td/div')  # 房間底注 / 籌碼 / 總輸贏 元素位置
textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newList1.remove('房間底注:')
newList1.remove('籌碼:')
newList1.remove('總輸贏:')     
            
newDataList1 = DataFrame(newList1)
newList1T = newDataList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv', header=True, index=False, encoding='utf-8-sig')

readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv')

newData = readCSV.rename(columns = {"0" : '房間底注', "1" : '籌碼', "2" : '總輸贏'})

csvFile1 = newData.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv', index=None, encoding='utf-8-sig')
print("第", 1, "筆資料寫入完成。", '\n')
time.sleep(1)


# =======================================================寫入第2~50筆測試紀錄(房間底注/籌碼/總輸贏)=======================================================

for i in range(2, 21):
    xpath_bet = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[1]/td/div'
    element_bet = driver.find_element(By.XPATH, xpath_bet)
    textSplit_bet = element_bet.text.split()
    newList_bet = list(textSplit_bet)
    newList_bet.remove('房間底注:')
    newList_bet.remove('籌碼:')
    newList_bet.remove('總輸贏:')  
    betlist = DataFrame(newList_bet)
    betlistT = betlist.T
    csvFile = betlistT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv', mode='a', header=False, index=False, encoding='utf-8-sig')
    print("第", i, "筆資料寫入完成。", '\n')

csvload = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv') 
excelFile1 = csvload.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.xlsx', header=True, index=False)       
time.sleep(1)

# =================================選取遊戲紀錄頁(2) =================================    
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[2]')
page_select.click()
time.sleep(3)

for i in range(1, 21):
    xpath_bet = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[1]/td/div'
    element_bet = driver.find_element(By.XPATH, xpath_bet)
    textSplit_bet = element_bet.text.split()
    newList_bet = list(textSplit_bet)
    newList_bet.remove('房間底注:')
    newList_bet.remove('籌碼:')
    newList_bet.remove('總輸贏:')  
    betlist = DataFrame(newList_bet)
    betlistT = betlist.T
    csvFile = betlistT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv', mode='a', header=False, index=False, encoding='utf-8-sig')
    print("第", i+20, "筆資料寫入完成。", '\n')

csvload = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv') 
excelFile1 = csvload.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.xlsx', header=True, index=False)       
time.sleep(1)

# =================================選取遊戲紀錄頁(3) =================================    
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[3]')
page_select.click()
time.sleep(3)

for i in range(1, 11):
    xpath_bet = f'//*[@id="content"]/div[{i}]/div/table/tbody/tr[1]/td/div'
    element_bet = driver.find_element(By.XPATH, xpath_bet)
    textSplit_bet = element_bet.text.split()
    newList_bet = list(textSplit_bet)
    newList_bet.remove('房間底注:')
    newList_bet.remove('籌碼:')
    newList_bet.remove('總輸贏:')  
    betlist = DataFrame(newList_bet)
    betlistT = betlist.T
    csvFile = betlistT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv', mode='a', header=False, index=False, encoding='utf-8-sig')
    print("第", i+40, "筆資料寫入完成。", '\n')

csvload = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.csv') 
excelFile1 = csvload.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.xlsx', header=True, index=False)       
time.sleep(1)


# ================================= 返回遊戲紀錄頁(1) =================================
# ===================== 寫入第1~20筆測試紀錄(中間區塊 '押注'/'贏分') =====================
  
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[1]')
page_select.click()
time.sleep(3)

xpath_M = f'//*[@id="content"]/div[1]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]'  # 中間區塊 '押注'/'贏分' 紀錄            
element_M = driver.find_element(By.XPATH, xpath_M)
textSplit_M = element_M.text.split()
newList_M = list(textSplit_M)
listX = [x for x in newList_M if (x != 'x2' and x != 'x3')]
newDataList_M = DataFrame(listX)
newList_MT = newDataList_M.T
csvFile = newList_MT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper1.csv', header=True, index=False, encoding='utf-8-sig')
print("第", 1, "筆資料寫入完成。", '\n')


for j in range(2, 21):
    xpath_M = f'//*[@id="content"]/div[{j}]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]'  # 中間區塊 '押注'/'贏分' 紀錄            
    element_M = driver.find_element(By.XPATH, xpath_M)
    textSplit_M = element_M.text.split()
    newList_M = list(textSplit_M)
    listX = [x for x in newList_M if (x != 'x2' and x != 'x3')]
    newDataList_M = DataFrame(listX)
    newList_MT = newDataList_M.T
    csvFile = newList_MT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper1.csv', mode='a', header=False, index=False, encoding='utf-8-sig')
    print("第", j, "筆資料寫入完成。", '\n')
    

time.sleep(1)


# =================================選取遊戲紀錄頁(2) ================================= 
# ======================寫入第21~40筆測試紀錄(中間區塊 '押注'/'贏分')======================  
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[2]')
page_select.click()
time.sleep(3)

for j in range(1, 21):
    xpath_M = f'//*[@id="content"]/div[{j}]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]'  # 中間區塊 '押注'/'贏分' 紀錄            
    element_M = driver.find_element(By.XPATH, xpath_M)
    textSplit_M = element_M.text.split()
    newList_M = list(textSplit_M)
    listX = [x for x in newList_M if (x != 'x2' and x != 'x3')]
    newDataList_M = DataFrame(listX)
    newList_MT = newDataList_M.T
    csvFile = newList_MT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper1.csv', mode='a', header=False, index=False, encoding='utf-8-sig')  
    print("第", j+20, "筆資料寫入完成。", '\n')

time.sleep(1)
            

# =================================選取遊戲紀錄頁(3) ================================= 
# ======================寫入第41~50筆測試紀錄(中間區塊 '押注'/'贏分')======================  
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[3]')
page_select.click()
time.sleep(3)  

for j in range(1, 11):
    xpath_M = f'//*[@id="content"]/div[{j}]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]'  # 中間區塊 '押注'/'贏分' 紀錄            
    element_M = driver.find_element(By.XPATH, xpath_M)
    textSplit_M = element_M.text.split()
    newList_M = list(textSplit_M)
    listX = [x for x in newList_M if (x != 'x2' and x != 'x3')]
    newDataList_M = DataFrame(listX)
    newList_MT = newDataList_M.T
    csvFile = newList_MT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper1.csv', mode='a', header=False, index=False, encoding='utf-8-sig')  
    print("第", j+40, "筆資料寫入完成。", '\n')
           
csvload = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper1.csv') 
excelFile2 = csvload.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper1.xlsx', header=False, index=False)        

time.sleep(1)

# ================================= 返回遊戲紀錄頁(1) =================================
# ==================== 寫入第1筆測試紀錄(下半區塊 '紀錄流水號~贏分') ====================
  
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[1]')
page_select.click()
time.sleep(3)

xpath0 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[2]/td')  # 遊戲紀錄流水號
xpath1 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[3]/td/span')  # 遊戲名稱
xpath2 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[6]/td')  # 玩家
xpath3 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[8]/td')  # 局號
xpath4 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[10]/td[1]')  # 結束時間
xpath5 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[10]/td[2]')  # 房間
xpath6 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[12]/td[1]')  # 序號
xpath7 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[12]/td[2]')  # 場景
xpath8 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[14]/td[1]')  # 面額
xpath9 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[14]/td[2]')  # 帳務
xpath10 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[14]/td[3]')  # 押注
xpath11 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[16]/td[1]')  # 彩金
xpath12 = driver.find_element(By.XPATH, '//*[@id="content"]/div[1]/div/table/tbody/tr[16]/td[2]')  # 贏分

textSplit0 = xpath0.text.split()
textSplit1 = xpath1.text.split()
textSplit2 = xpath2.text.split()
textSplit3 = xpath3.text.split()
textSplit4 = xpath4.text.split()
textSplit5 = xpath5.text.split()
textSplit6 = xpath6.text.split()
textSplit7 = xpath7.text.split()
textSplit8 = xpath8.text.split()
textSplit9 = xpath9.text.split()
textSplit10 = xpath10.text.split()
textSplit11 = xpath11.text.split()
textSplit12 = xpath12.text.split()

newList0 = list(textSplit0)
newList1 = list(textSplit1)
newList2 = list(textSplit2)
newList3 = list(textSplit3)
newList4 = list(textSplit4)
newList4_1 = ''.join(' ').join(newList4)
newList4_2 = []
newList4_2.append(newList4_1)
newList5 = list(textSplit5)
newList6 = list(textSplit6)
newList7 = list(textSplit7)
newList8 = list(textSplit8)
newList9 = list(textSplit9)
newList10 = list(textSplit10)
newList11 = list(textSplit11)
newList12 = list(textSplit12)

totalList = []
totalList.extend([newList0, newList1, newList2, newList3, newList4_2, newList5, newList6, newList7, newList8, newList9, newList10, newList11, newList12])

totalListDT = DataFrame(totalList)
totalListDTT = totalListDT.T

csvFileNT = totalListDT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_NT.csv', index=False, encoding='utf-8-sig')
csvFile = totalListDTT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', index=False, encoding='utf-8-sig')


readData = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv')
newData = readData.rename(columns = {'0':'紀錄流水號', '1':'遊戲名稱', '2':'玩家', '3':'局號', '4':'結束時間',
                                     '5':'房間', '6':'序號', '7':'場景', '8':'面額', '9':'帳務',
                                     '10':'押注', '11':'彩金', '12':'贏分'})
newFile1 = newData.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', index=False, encoding = 'utf-8-sig')

excelFileNT = totalListDT.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_NT.xlsx', index=None, header=True)
print('第', 1, '筆資料寫入完成。', '\n')


# =======================================================寫入第2~20筆測試紀錄(下半區塊 '紀錄流水號~贏分')=======================================================

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[2]/td'  # 紀錄流水號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')      

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[3]/td/span'  # 遊戲名稱
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[6]/td'  # 玩家名稱
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[8]/td'  # 局號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[10]/td[1]'  # 結束時間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList_1 = ''.join(' ').join(newDataList)
    newDataList_2 = []
    newDataList_2.append(newDataList_1)
    newDataList = DataFrame(newDataList_2)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[10]/td[2]'  # 房間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[12]/td[1]'  # 序號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[12]/td[2]'  # 場景
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[1]'  # 面額
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[2]'  # 帳務
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[3]'  # 押注
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[16]/td[1]'  # 彩金
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(2, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[16]/td[2]'  # 贏分 
    element = driver.find_element(By.XPATH, xpath) 
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    print('第', k, '筆資料寫入完成。', '\n')


# =================================選取遊戲紀錄頁(2) ================================= 
# ======================寫入第21~40筆測試紀錄(下半區塊 '紀錄流水號~贏分')======================  
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[2]')
page_select.click()
time.sleep(3)


for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[2]/td'  # 紀錄流水號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')      

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[3]/td/span'  # 遊戲名稱
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[6]/td'  # 玩家名稱
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[8]/td'  # 局號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[10]/td[1]'  # 結束時間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList_1 = ''.join(' ').join(newDataList)
    newDataList_2 = []
    newDataList_2.append(newDataList_1)
    newDataList = DataFrame(newDataList_2)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[10]/td[2]'  # 房間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[12]/td[1]'  # 序號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[12]/td[2]'  # 場景
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[1]'  # 面額
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[2]'  # 帳務
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[3]'  # 押注
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[16]/td[1]'  # 彩金
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 21):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[16]/td[2]'  # 贏分 
    element = driver.find_element(By.XPATH, xpath) 
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    print('第', k+20, '筆資料寫入完成。', '\n')
    

# =================================選取遊戲紀錄頁(3) ================================= 
# ======================寫入第41~50筆測試紀錄(下半區塊 '紀錄流水號~贏分')======================  
    
listPage = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="page_selector"]')))
actions = ActionChains(driver)
actions.move_to_element(listPage)    
actions.perform()
listPage.click()
time.sleep(2)

page_select = driver.find_element(By.XPATH, '//*[@id="page_selector"]/option[3]')
page_select.click()
time.sleep(3)


for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[2]/td'  # 紀錄流水號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')      

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[3]/td/span'  # 遊戲名稱
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[6]/td'  # 玩家名稱
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[8]/td'  # 局號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[10]/td[1]'  # 結束時間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList_1 = ''.join(' ').join(newDataList)
    newDataList_2 = []
    newDataList_2.append(newDataList_1)
    newDataList = DataFrame(newDataList_2)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[10]/td[2]'  # 房間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[12]/td[1]'  # 序號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[12]/td[2]'  # 場景
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[1]'  # 面額
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[2]'  # 帳務
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[14]/td[3]'  # 押注
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[16]/td[1]'  # 彩金
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for k in range(1, 11):
    xpath = f'//*[@id="content"]/div[{k}]/div/table/tbody/tr[16]/td[2]'  # 贏分 
    element = driver.find_element(By.XPATH, xpath) 
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    print('第', k+40, '筆資料寫入完成。', '\n')

readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.csv')
writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.xlsx', header=True, index=False)

time.sleep(2)
driver.close()

workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.xlsx')
source_sheet = workbook['Sheet1']

source_sheet.move_range('A22' ":" 'A40', rows = -19, cols = 1)
source_sheet.move_range('A41' ":" 'A59', rows = -38, cols = 2)
source_sheet.move_range('A60' ":" 'A78', rows = -57, cols = 3)
source_sheet.move_range('A79' ":" 'A97', rows = -76, cols = 4)
source_sheet.move_range('A98' ":" 'A116', rows = -95, cols = 5)
source_sheet.move_range('A117' ":" 'A135', rows = -114, cols = 6)
source_sheet.move_range('A136' ":" 'A154', rows = -133, cols = 7)
source_sheet.move_range('A155' ":" 'A173', rows = -152, cols = 8)
source_sheet.move_range('A174' ":" 'A192', rows = -171, cols = 9)
source_sheet.move_range('A193' ":" 'A211', rows = -190, cols = 10)
source_sheet.move_range('A212' ":" 'A230', rows = -209, cols = 11)
source_sheet.move_range('A231' ":" 'A249', rows = -228, cols = 12)
source_sheet.move_range('A250' ":" 'A269', rows = -228, cols = 0)
source_sheet.move_range('A270' ":" 'A289', rows = -248, cols = 1)
source_sheet.move_range('A290' ":" 'A309', rows = -268, cols = 2)
source_sheet.move_range('A310' ":" 'A329', rows = -288, cols = 3)
source_sheet.move_range('A330' ":" 'A349', rows = -308, cols = 4)
source_sheet.move_range('A350' ":" 'A369', rows = -328, cols = 5)
source_sheet.move_range('A370' ":" 'A389', rows = -348, cols = 6)
source_sheet.move_range('A390' ":" 'A409', rows = -368, cols = 7)
source_sheet.move_range('A410' ":" 'A429', rows = -388, cols = 8)
source_sheet.move_range('A430' ":" 'A449', rows = -408, cols = 9)
source_sheet.move_range('A450' ":" 'A469', rows = -428, cols = 10)
source_sheet.move_range('A470' ":" 'A489', rows = -448, cols = 11)
source_sheet.move_range('A490' ":" 'A509', rows = -468, cols = 12)
source_sheet.move_range('A510' ":" 'A519', rows = -468, cols = 0)
source_sheet.move_range('A520' ":" 'A529', rows = -478, cols = 1)
source_sheet.move_range('A530' ":" 'A539', rows = -488, cols = 2)
source_sheet.move_range('A540' ":" 'A549', rows = -498, cols = 3)
source_sheet.move_range('A550' ":" 'A559', rows = -508, cols = 4)
source_sheet.move_range('A560' ":" 'A569', rows = -518, cols = 5)
source_sheet.move_range('A570' ":" 'A579', rows = -528, cols = 6)
source_sheet.move_range('A580' ":" 'A589', rows = -538, cols = 7)
source_sheet.move_range('A590' ":" 'A599', rows = -548, cols = 8)
source_sheet.move_range('A600' ":" 'A609', rows = -558, cols = 9)
source_sheet.move_range('A610' ":" 'A619', rows = -568, cols = 10)
source_sheet.move_range('A620' ":" 'A629', rows = -578, cols = 11)
source_sheet.move_range('A630' ":" 'A639', rows = -588, cols = 12)

 
workbook.save(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower_v1.xlsx')
workbook.close()

print("EXCEL檔案儲存成功!", '\n')

time.sleep(2)
print('(菲律賓骰寶)前台共50筆遊戲紀錄獲取完成!', '\n')


# ============================Section.1 進入後台首頁============================

# ch_options = Options()
# ch_options.add_argument("--headless")  #無WEB UI顯示
# driver = webdriver.Chrome(ch_options)

driver = webdriver.Chrome()  # 有WEB UI顯示

time.sleep(1) 
driver.get("https://dev-admin.claretfox.com/")
http_status = requests.get('https://dev-admin.claretfox.com/')
stateCode = http_status.status_code
print('Http Response Code: ', stateCode, '\n')
if stateCode == 200:
    print("HTTP回應成功!", '\n')
        
else:
    print('Http Response Code:', stateCode, '\n')
    print("HTTP回應失敗!", '\n')
    driver.quit()
    print('自動化測試已中斷!', '\n')
time.sleep(2)

logging.basicConfig(level=logging.DEBUG,
                    filename='output.log',
                    datefmt='%Y/%m/%d %H:%M:%S',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(lineno)d - %(module)s - %(message)s')
logger = logging.getLogger(__name__)

logger.info('This is a log info')
logger.debug('Debugging')
logger.warning('Warning exists')
logger.info('Finish')



back_platform = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div/div/div/form/div/div[1]")))
back_platform.click()
time.sleep(2)
print("進入DEV後台首頁!", '\n')

# -----------切換網頁顯示語系-----------
language_bar = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/form/label')))
language_bar.click()

language_ch_zh = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="lang"]/option[2]')))
language_ch_zh.click()
print("語系已切換'繁體中文'!", '\n')

# ============================Section.2 登入使用者名稱及密碼============================

login_icon = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, 'content-group')))
login_icon.click()  # 尋找登入介面元素位址
time.sleep(2)

login_acc = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="user_id"]')))
login_acc.clear()    # 預設此欄位為null, 但仍先清除帳號欄位資訊
login_pass = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]')))   
login_pass.clear()   # 預設此欄位為null, 但仍先清除密碼欄位資訊

login_verificationCode = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[4]/input')))  # 驗證碼欄位

verificationCode = driver.find_element(By.CSS_SELECTOR, '#captcha_img > img')
verificationCode.screenshot('verificationCode.png')
time.sleep(1)

actions = ActionChains(driver)
actions.move_to_element(login_acc)    
actions.perform()
login_acc.send_keys("ivan_li")   # 個人使用者帳號
time.sleep(1)
login_acc.send_keys(Keys.TAB)   # 切換至密碼輸入欄位
login_pass.send_keys("iPlaystar296")  # 個人密碼
time.sleep(1)
login_pass.send_keys(Keys.TAB)   # 切換至驗證碼輸入欄位
ocr = ddddocr.DdddOcr()
with open('verificationCode.png', 'rb') as fp:
    image = fp.read()
catch = ocr.classification(image)
login_verificationCode.send_keys(catch)
time.sleep(1)

login_button = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/form/div/div[6]/button')))
login_button.click()
print("登入成功!", '\n')
time.sleep(2)

# ============================Section.3 切換後台功能頁籤============================

player = driver.find_element(By.ID, 'Player')
player.click()
print("進入玩家功能選單!", '\n')
time.sleep(2)

gameRecord = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/div/div/ul/li[3]/ul/li[6]/a')
actions.move_to_element(gameRecord)    
actions.perform()    
gameRecord.click()
print("切換遊戲紀錄子選單!", '\n')
time.sleep(2)

start_time = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div')))

startTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[2]/div/div/span/button')
actions.move_to_element(startTime_button)    
actions.perform()
startTime_button.click()

# startDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
# actions.move_to_element(startDate)    
# actions.perform()
# startDate.click()

startTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[4]/a')  # 開始日期選取 "2024.10.24"
startTime_check.click()
time.sleep(1)

startTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
startTime_confirm.click()
time.sleep(2)

endTime_button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[4]/div[2]/div[1]/div/div/div[2]/div[3]/div/div/span/button/i')
actions.move_to_element(endTime_button)    
actions.perform()
endTime_button.click()

# endDate = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[1]/a[1]/span')  # 選取 "<"鍵跳至前月日期選單
# actions.move_to_element(endDate)    
# actions.perform()
# endDate.click()

endTime_check = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/table/tbody/tr[4]/td[4]/a')  # 結束日期選取 "2024.10.24"
endTime_check.click()
time.sleep(1)

endTime_confirm = driver.find_element(By.XPATH, '//*[@id="ui-datepicker-div"]/div[3]/button[2]')
endTime_confirm.click()
time.sleep(2)

showLimit = driver.find_element(By.XPATH, '//*[@id="count"]')
showLimit.click()
time.sleep(2)

showLimitSet = driver.find_element(By.XPATH, '//*[@id="count"]/option[2]')  # 設定顯示筆數 = 50筆
showLimitSet.click()
time.sleep(2)

agencyType = driver.find_element(By.XPATH, '//*[@id="agent_attr"]')
agencyType.click()
time.sleep(2)

agencyTypeSet = driver.find_element(By.XPATH, '//*[@id="agent_attr"]/option[10]')  # 設定代理商類別 = Test
agencyTypeSet.click()
time.sleep(2)
print('代理商類別: ', agencyTypeSet.text, '\n')

agencyName = driver.find_element(By.XPATH, '//*[@id="agent"]')
agencyName.click()
time.sleep(2)

agencyNameSet = driver.find_element(By.XPATH, '//*[@id="agent"]/option[8]')  # 設定代理商名稱 = Test-2
agencyNameSet.click()
time.sleep(2)
print('代理商名稱: ', agencyNameSet.text, '\n')

gameType = driver.find_element(By.XPATH, '//*[@id="game_type"]')
gameType.click()
time.sleep(2)

gameTypeSet = driver.find_element(By.XPATH, '//*[@id="game_type"]/option[7]')  # 設定遊戲類別 = 棋牌遊戲
driver.execute_script("arguments[0].scrollIntoView()", gameTypeSet)
gameTypeSet.click()
time.sleep(2)
print('遊戲類別: ', gameTypeSet.text, '\n')

gameSelect = driver.find_element(By.XPATH, '//*[@id="game_select"]')
gameSelect.click()
time.sleep(2)

gameSelectSet = driver.find_element(By.XPATH, '//*[@id="game_select"]/option[8]')  # 遊戲選擇 = 菲律賓骰寶
gameSelectSet.click()
time.sleep(2)
print('遊戲名稱: ', gameSelectSet.text, '\n')

playerName = driver.find_element(By.XPATH, '//*[@id="player"]')  # 玩家名稱 = 'ivan_li'
playerName.send_keys("ivan_li")
time.sleep(2)

btnSubmit = driver.find_element(By.XPATH, '//*[@id="sh_btn"]')  # 確認無誤送出
btnSubmit.click()
time.sleep(10)

print('獲取(菲律賓骰寶)後台遊戲紀錄中...', '\n')


path_1 = r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data'
if not os.path.isdir(path_1):
    os.makedirs(path_1)
    

# =======================================================寫入第1筆測試紀錄(房間底注/籌碼/總輸贏)=======================================================

xpath1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[1]/td/div')  # 房間底注 / 籌碼 / 總輸贏 元素位置
textSplit1 = xpath1.text.split()
newList1 = list(textSplit1)
newList1.remove('房間底注:')
newList1.remove('籌碼:')
newList1.remove('總輸贏:')     
            
newDataList1 = DataFrame(newList1)
newList1T = newDataList1.T
csvFile = newList1T.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper0.csv', header=True, index=False, encoding='utf-8-sig')

readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper0.csv')

newData = readCSV.rename(columns = {"0" : '房間底注', "1" : '籌碼', "2" : '總輸贏'})

csvFile1 = newData.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper0.csv', index=None, encoding='utf-8-sig')
print("第", 1, "筆資料寫入完成。", '\n')
time.sleep(1)


# =======================================================寫入第2~50筆測試紀錄(房間底注/籌碼/總輸贏)=======================================================

for i in range(2, 51):
    xpath_bet = f' //*[@id="history"]/div[{i}]/div/table/tbody/tr[1]/td/div'
    element_bet = driver.find_element(By.XPATH, xpath_bet)
    textSplit_bet = element_bet.text.split()
    newList_bet = list(textSplit_bet)
    newList_bet.remove('房間底注:')
    newList_bet.remove('籌碼:')
    newList_bet.remove('總輸贏:')  
    betlist = DataFrame(newList_bet)
    betlistT = betlist.T
    csvFile = betlistT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper0.csv', mode='a', header=False, index=False, encoding='utf-8-sig')
    print("第", i, "筆資料寫入完成。", '\n')

csvload = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper0.csv') 
excelFile1 = csvload.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper0.xlsx', header=True, index=False)       
time.sleep(1)


# =======================================================寫入第1~50筆測試紀錄(中間區塊 '押注'/'贏分')=======================================================

xpath_M = f'//*[@id="history"]/div[1]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]'  # 中間區塊 '押注'/'贏分' 紀錄
element_M = driver.find_element(By.XPATH, xpath_M)
textSplit_M = element_M.text.split()
newList_M = list(textSplit_M)
listX = [x for x in newList_M if (x != 'x2' and x != 'x3')]
newDataList_M = DataFrame(listX)
newList_MT = newDataList_M.T
csvFile = newList_MT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper1.csv', header=True, index=False, encoding='utf-8-sig')

for j in range(2, 51):
    xpath_M = f'//*[@id="history"]/div[{j}]/div/table/tbody/tr[4]/td/div/div/div/div/div[2]'  # 中間區塊 '押注'/'贏分' 紀錄
    element_M = driver.find_element(By.XPATH, xpath_M)
    textSplit_M = element_M.text.split()
    newList_M = list(textSplit_M)
    listX = [x for x in newList_M if (x != 'x2' and x != 'x3')]
    newDataList_M = DataFrame(listX)
    newList_MT = newDataList_M.T
    csvFile = newList_MT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper1.csv', mode='a', header=False, index=False, encoding='utf-8-sig')
        
csvload = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper1.csv') 
excelFile2 = csvload.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper1.xlsx', header=False, index=False)        
time.sleep(1)       


# =======================================================寫入第1筆測試紀錄(紀錄流水號~贏分)=======================================================

xpath0 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[2]/td')  # 遊戲紀錄流水號
xpath1 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[3]/td')  # 遊戲名稱
xpath2 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[6]/td')  # 玩家
xpath3 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[8]/td')  # 局號
xpath4 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[10]/td[2]')  # 結束時間
xpath5 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[12]/td[3]')  # 房間
xpath6 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[14]/td[1]')  # 序號
xpath7 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[12]/td[2]')  # 場景
xpath8 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[14]/td[2]')  # 面額
xpath9 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[14]/td[3]')  # 帳務
xpath10 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[14]/td[4]')  # 押注
xpath11 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[16]/td[1]')  # 彩金
xpath12 = driver.find_element(By.XPATH, '//*[@id="history"]/div[1]/div/table/tbody/tr[16]/td[2]')  # 贏分

textSplit0 = xpath0.text.split()
textSplit1 = xpath1.text[12:].split()
textSplit2 = xpath2.text[9:].split()
textSplit3 = xpath3.text.split()
textSplit4 = xpath4.text.split()
textSplit5 = xpath5.text.split()
textSplit6 = xpath6.text.split()
textSplit7 = xpath7.text.split()
textSplit8 = xpath8.text.split()
textSplit9 = xpath9.text.split()
textSplit10 = xpath10.text.split()
textSplit11 = xpath11.text.split()
textSplit12 = xpath12.text.split()

newList0 = list(textSplit0)
newList1 = list(textSplit1)
newList2 = list(textSplit2)
newList3 = list(textSplit3)
newList4 = list(textSplit4)
newList4_1 = ''.join(' ').join(newList4)
newList4_2 = []
newList4_2.append(newList4_1)
newList5 = list(textSplit5)
newList6 = list(textSplit6)
newList7 = list(textSplit7)
newList8 = list(textSplit8)
newList9 = list(textSplit9)
newList10 = list(textSplit10)
newList11 = list(textSplit11)
newList12 = list(textSplit12)


totalList = []
totalList.extend([newList0, newList1, newList2, newList3, newList4_2, newList5, newList6, newList7, newList8, newList9, newList10, newList11, newList12])
# print(totalList, '\n')

totalListDT = DataFrame(totalList)
totalListDTT = totalListDT.T

csvFileNT = totalListDT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_NT.csv', index=0, encoding='utf-8-sig')
csvFile = totalListDTT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', index=False, encoding='utf-8-sig')


readData = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv')
newData = readData.rename(columns = {'0':'紀錄流水號', '1':'遊戲名稱', '2':'玩家', '3':'局號', '4':'結束時間',
                                     '5':'房間', '6':'序號', '7':'場景', '8':'面額', '9':'帳務',
                                     '10':'押注', '11':'彩金', '12':'贏分'})
newFile1 = newData.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', index=False, encoding = 'utf-8-sig')

excelFileNT = totalListDT.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_NT.xlsx', index=None, header=True)
print('第', 1, '筆資料寫入完成。', '\n')


for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[2]/td'  # 紀錄流水號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[3]/td'  # 遊戲名稱
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text[12:].split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[6]/td'  # 玩家名稱
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text[9:].split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[8]/td'  # 局號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[10]/td[2]'  # 結束時間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList_1 = ''.join(' ').join(newDataList)
    newDataList_2 = []
    newDataList_2.append(newDataList_1)
    newDataList = DataFrame(newDataList_2)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[12]/td[3]'  # 房間
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[14]/td[1]'  # 序號
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[12]/td[2]'  # 場景
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[14]/td[2]'  # 面額
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[14]/td[3]'  # 帳務
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[14]/td[4]'  # 押注
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[16]/td[1]'  # 彩金
    element = driver.find_element(By.XPATH, xpath)
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')

for i in range(2, 51):
    xpath = f'//*[@id="history"]/div[{i}]/div/table/tbody/tr[16]/td[2]'  # 贏分 
    element = driver.find_element(By.XPATH, xpath) 
    dataSplit = element.text.split()
    newDataList = list(dataSplit)
    newDataList = DataFrame(newDataList)
    newDataListT = newDataList.T
    totalList = newDataListT.to_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv', mode='a', header=None, index=False, encoding='utf-8-sig')
    readCSV = pd.read_csv(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.csv')
    writeEX = readCSV.to_excel(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.xlsx', header=True, index=False)
    print('第', i, '筆資料寫入完成。', '\n')
    
    
workbook = openpyxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.xlsx')
source_sheet = workbook['Sheet1']

source_sheet.move_range('A52' ":" 'A100', rows = -49, cols = 1)
source_sheet.move_range('A101' ":" 'A149', rows = -98, cols = 2)
source_sheet.move_range('A150' ":" 'A198', rows = -147, cols = 3)
source_sheet.move_range('A199' ":" 'A247', rows = -196, cols = 4)
source_sheet.move_range('A248' ":" 'A296', rows = -245, cols = 5)
source_sheet.move_range('A297' ":" 'A345', rows = -294, cols = 6)
source_sheet.move_range('A346' ":" 'A394', rows = -343, cols = 7)
source_sheet.move_range('A395' ":" 'A443', rows = -392, cols = 8)
source_sheet.move_range('A444' ":" 'A492', rows = -441, cols = 9)
source_sheet.move_range('A493' ":" 'A541', rows = -490, cols = 10)
source_sheet.move_range('A542' ":" 'A590', rows = -539, cols = 11)
source_sheet.move_range('A591' ":" 'A639', rows = -588, cols = 12)
 
workbook.save(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower_v1.xlsx')
workbook.close()

print("EXCEL檔案儲存成功!", '\n')

time.sleep(2)
driver.close()
print('(菲律賓骰寶)後台共50筆遊戲紀錄獲取完成!', '\n')
   

# =================================================== Step.3 前/後台資料分析 ===================================================

# print('前/後台遊戲紀錄分析中...', '\n')

path_01 = r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform'
if not os.path.isdir(path_01):
    os.makedirs(path_01)
    
path_02 = r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform'
if not os.path.isdir(path_02):
    os.makedirs(path_02)
    
path_03 = r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge'
if not os.path.isdir(path_03):
    os.makedirs(path_03) 


# ============================= (上半部 : 房間底注 / 籌碼 / 總輸贏) =============================

# 1.數據比對 

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper0.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_upper0_analysis.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_upper0_analysis.xlsx')
workbook_2.close()
time.sleep(1)
print('前/後台遊戲紀錄比對完成!', '\n')

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper0.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper0.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_upper0_analysis.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_upper0_analysis.xlsx')

print('前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(upper0).xlsx")
print("前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(upper0).xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '前台原始遊戲紀錄(菲律賓骰寶)_20241024'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '後台原始遊戲紀錄(菲律賓骰寶)_20241024'  # 修改分頁.2工作表名稱
ws_2 = wb['Sheet1_2']
ws_2.title = '前台遊戲紀錄比對結果(菲律賓骰寶)_20241024'  # 修改分頁.3工作表名稱
ws_3 = wb['Sheet1_3']
ws_3.title = '後台遊戲紀錄比對結果(菲律賓骰寶)_20241024'  # 修改分頁.4工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(upper0).xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
sheet_1 = wb.worksheets[2]  # 分頁.3
sheet_1.sheet_properties.tabColor = 'CC6600'
sheet_2 = wb.worksheets[3]  # 分頁.3
sheet_2.sheet_properties.tabColor = 'AAFFEE'

wb.save(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(upper0).xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()


# ============================= (中間段 : 押注 / 贏分) =============================

# 1.數據比對 (中間段 : 押注 / 贏分)

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper1.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper1.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_upper1_analysis.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_upper1_analysis.xlsx')
workbook_2.close()
time.sleep(1)
print('前/後台遊戲紀錄比對完成!', '\n')

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_upper1.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_upper1.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_upper1_analysis.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_upper1_analysis.xlsx')

print('前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(upper1).xlsx")
print("前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(upper1).xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '前台原始遊戲紀錄(菲律賓骰寶)_20241024'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '後台原始遊戲紀錄(菲律賓骰寶)_20241024'  # 修改分頁.2工作表名稱
ws_2 = wb['Sheet1_2']
ws_2.title = '前台遊戲紀錄比對結果(菲律賓骰寶)_20241024'  # 修改分頁.3工作表名稱
ws_3 = wb['Sheet1_3']
ws_3.title = '後台遊戲紀錄比對結果(菲律賓骰寶)_20241024'  # 修改分頁.4工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(upper1).xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
sheet_1 = wb.worksheets[2]  # 分頁.3
sheet_1.sheet_properties.tabColor = 'CC6600'
sheet_2 = wb.worksheets[3]  # 分頁.3
sheet_2.sheet_properties.tabColor = 'AAFFEE'

wb.save(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(upper1).xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()


# =========== (下半部 : 玩家 / 局號 / 結束時間 / 場景 / 房間 / 序號 / 面額 / 帳務 / 押注 / 彩金 / 贏分) ===========

# 1.數據比對 (下半部 : 玩家 / 局號 / 結束時間 / 場景 / 房間 / 序號 / 面額 / 帳務 / 押注 / 彩金 / 贏分)

# 下面使用openpyxl模块的load_workbook函数读取到Excel文件对象，并提取两个Excel文件中'Sheet1'工作表作为源数据。

workbook_1 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower_v1.xlsx')

workbook_2 = pxl.load_workbook(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower_v1.xlsx')

# 读取表1的sheet1内容

workbook_1_sheet_1 = workbook_1['Sheet1']

# 读取表2的sheet1内容

workbook_2_sheet_1 = workbook_2['Sheet1']

# 提取两个工作表中的最大行和最大列，这样即使两个表的行数和列数不一致也能完全找出不同的单元格数据。
# 确定最大行
if workbook_1_sheet_1.max_row > workbook_2_sheet_1.max_row:
    max_row = workbook_1_sheet_1.max_row
else:
    max_row = workbook_2_sheet_1.max_row
# 确定最大列
if workbook_1_sheet_1.max_column > workbook_2_sheet_1.max_column:
    max_column = workbook_1_sheet_1.max_column
else:
    max_column = workbook_2_sheet_1.max_column

# 使用for循环的方式分别遍历行数据和列数据，然后判断对应单元格的数据值是否相等，若是不相等则打上标记。

for i in range(1, (max_row + 1)):

    for j in range(1, (max_column + 1)):

        cell_1 = workbook_1_sheet_1.cell(i, j)

        cell_2 = workbook_2_sheet_1.cell(i, j)

        if cell_1.value != cell_2.value:

            cell_1.fill = PatternFill("solid", fgColor='FFFF00')

            cell_1.font = Font(color=colors.BLACK, bold=False)  # 粗體字 "bold=True"

            cell_2.fill = PatternFill("solid", fgColor='FFBB00')

            cell_2.font = Font(color=colors.BLACK, bold=False)


# 将 前/後台有差異欄位標記"黃底"並匯出
# print('前台比對資料匯出中...', '\n')
workbook_1.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_lower_analysis.xlsx')
workbook_1.close()
time.sleep(1)

# print('後台比對資料匯出中...', '\n')
workbook_2.save(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_lower_analysis.xlsx')
workbook_2.close()
time.sleep(1)
print('前/後台遊戲紀錄比對完成!', '\n')

# 创建一个列表，并将需要合并的Excel文件放入其中
files = []
files.append(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower_v1.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower_v1.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_lower_analysis.xlsx')
files.append(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_lower_analysis.xlsx')

print('前/後台比對資料合併中...', '\n') 
# 创建一个新工作簿
newbook = Workbook()
# 删除其中的默认工作表
newbook.Worksheets.Clear()
 
# 创建一个临时工作簿
tempbook = Workbook()
# 遍历列表中的文件路径
for file in files:
    # 将列表中的文件加载到临时工作簿中
    tempbook.LoadFromFile(file)
    # 遍历临时工作簿中所有的工作表
    for sheet in tempbook.Worksheets:
        # 将临时工作簿中的工作表复制到新工作簿中
        newbook.Worksheets.AddCopy(sheet, WorksheetCopyType.CopyAll)
 
# 将新工作簿另存为.xlsx文件
newbook.SaveToFile(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(lower).xlsx")
print("前/後比對資料合併完成!", '\n')
newbook.Dispose()
tempbook.Dispose()

# 2.修改分頁名稱
filename = r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(lower).xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']
ws.title = '前台原始遊戲紀錄(菲律賓骰寶)_20241024'  # 修改分頁.1工作表名稱
ws_1 = wb['Sheet1_1']
ws_1.title = '後台原始遊戲紀錄(菲律賓骰寶)_20241024'  # 修改分頁.2工作表名稱
ws_2 = wb['Sheet1_2']
ws_2.title = '前台遊戲紀錄比對結果(菲律賓骰寶)_20241024'  # 修改分頁.3工作表名稱
ws_3 = wb['Sheet1_3']
ws_3.title = '後台遊戲紀錄比對結果(菲律賓骰寶)_20241024'  # 修改分頁.4工作表名稱

sheetName = 'Evaluation Warning'
del wb[sheetName]  # 刪除多餘分頁，名稱='Evaluation Warning'之分頁
wb.save(filename)  # 儲存變更
print('工作表格式修改中...', '\n')
# 3.工作表顏色設定
wb = openpyxl.load_workbook(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(lower).xlsx")
sheet_1 = wb.worksheets[0]  # 分頁.1
sheet_1.sheet_properties.tabColor = '00BFFF'
sheet_2 = wb.worksheets[1]  # 分頁.2
sheet_2.sheet_properties.tabColor = '00FA9A'
sheet_1 = wb.worksheets[2]  # 分頁.3
sheet_1.sheet_properties.tabColor = 'CC6600'
sheet_2 = wb.worksheets[3]  # 分頁.3
sheet_2.sheet_properties.tabColor = 'AAFFEE'

wb.save(r"C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(3). Data_Merge\Data Merge_20241024(lower).xlsx")
print("工作表格式修改完成!", '\n')
time.sleep(1)
wb.close()


# ================================== 刪除資料分析後多餘檔案 ==================================

os.remove(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_NT.csv')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_NT.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(1). Front_Platform\(2). Test_Report\Raw_Data\front-platform_record(Color-Gmae_20241024)_lower.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_NT.csv')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_NT.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(2). Back_Platform\(2). Test_Report\Raw_Data\back-platform_record(Color-Gmae_20241024)_lower.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_lower_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_upper0_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(1). Front-platform\front-platform_record(Color-Gmae_20241024)_upper1_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_lower_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_upper0_analysis.xlsx')
os.remove(r'C:\AutomotiveTest\(3). Color Game\(3). Data Analysis\(2). Back-platform\back-platform_record(Color-Gmae_20241024)_upper1_analysis.xlsx')


       
time.sleep(1)
end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
print('測試結束時間: ', end_time, '\n')
    
